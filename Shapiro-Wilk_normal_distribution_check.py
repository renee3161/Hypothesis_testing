# -*- coding: utf-8 -*-
"""
Created on Thu Feb  6 09:37:12 2025

@author: rwats
"""

import pandas as pd
import openpyxl
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from scipy import stats
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO

# Read the data
file_path = 'Customer_sales_formatted.xlsx'
df = pd.read_excel(file_path)

# Set up the new Excel file to store results
wb = Workbook()
ws = wb.active
ws.title = "Sales Data Analysis"

# Write headers
headers = ['Customer', 'Shapiro-Wilk p-value', 'Normality']
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)
    
    
ws_plots = wb.create_sheet(title="KDE Plots")

# Adjust column widths and row heights for the Plots sheet to fit the plots
ws_plots.column_dimensions['A'].width = 50  # Set width for customer name column

# Process the data for each customer and perform the Shapiro-Wilk test
row_num = 2  # Start writing from row 2 (since row 1 is for headers)

for _, row in df.iterrows():
    customer_name = row['Customer']
    sales_data = row[1:].dropna()  # Drop any NaN values (missing sales data)

    #Perform the Shapiro-Wilk test
    if len(sales_data) >= 3:  # The test requires at least 3 data points
        stat, p_value = stats.shapiro(sales_data)

        # Determine normality
        normality = 'Yes' if p_value > 0.05 else 'No'   
        
        # Write results to the Excel sheet
        ws.cell(row=row_num, column=1, value=customer_name)
        ws.cell(row=row_num, column=2, value=p_value)
        ws.cell(row=row_num, column=3, value=normality)
        
       # Plot the Probability Density Function (KDE) for the customer sales data
        plt.figure(figsize=(6, 4))
       
       # Create the KDE plot using seaborn
        sns.kdeplot(sales_data, color='g', alpha=0.6)
    
       # Add title and labels
        plt.title(f"KDE of Sales Data for {customer_name}")
        plt.xlabel('Sales')
        plt.ylabel('Density')
    
       # Save the graph to the Excel file (no file system saving)
        image_stream = BytesIO()
        plt.savefig(image_stream, format='png')
        plt.close()  # Close the plot to release memory
    
        image_stream.seek(0)  # Rewind the stream to the beginning
        img = Image(image_stream)
       
        # Insert the image into the Excel sheet
        row_pos = (row_num - 2) * 10 + 1  # Adjusting where the plot starts, leave space between each plot
        img.width = 400  # Resize image to fit in the cell
        img.height = 500
        ws_plots.add_image(img, f'A{row_pos}')  # Place the image starting from column A
    
       # Increment row for the next customer
        row_num += 1

#Save the workbook to an Excel file
wb.save('customer_sales_analysis_with_kde.xlsx')

print("Excel file created with analysis and KDE graphs.")
