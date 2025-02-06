# -*- coding: utf-8 -*-
"""
Created on Thu Feb  6 09:48:41 2025

@author: rwats
"""

import pandas as pd
import openpyxl
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from scipy import stats
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
from collections import defaultdict


# take the file in format where each sales data is below the customer
# in the same column and change the file layout where each row
# contains a customer and the columns are sales

# Read the Excel file
file_path = 'customer_sales_2020_2024.xlsx'
df = pd.read_excel(file_path, header=None)

# Initialize variables
customer_names = []  # List to store customer names
sales_data = []  # List to store sales data for each customer across months
sales = {}
row_num = 2

wb = openpyxl.Workbook()
ws = wb.active  # Get the active sheet


columns = ['Customer'] + [f'{month} {year}' for year in range(2020, 2025) for month in ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']]
ws.append(columns)  # Add headers to the first row

count = 0
# Loop through the rows to process the data
current_customer = None
for index, row in df.iterrows():
    # Check if the row contains a customer name
    if count <= 35: # store the first customer sales data
        cell_value = row[0]
        converted_date = pd.to_datetime(cell_value, errors='coerce')
        is_valid_date = not pd.isna(converted_date)
        if not is_valid_date:
            sales['Customer'] = row[0]
        else:
            sales[row[0]] = row[1]
    count +=1
    if count > 35: # store all other customers sale data
        cell_value = row[0]
        converted_date = pd.to_datetime(cell_value, errors='coerce')
        is_valid_date = not pd.isna(converted_date)
        if not is_valid_date:
            for month in columns:  # Start from index 1 to skip the 'Customer' header
                if month in sales:
                    # Find the column index for this month
                    col_index = columns.index(month) + 1  # +1 because openpyxl is 1-indexed
                    # Write the sales value to the correct column for each month
                    ws.cell(row=row_num, column=col_index, value=sales[month])
            sales = {}
            sales['Customer'] = row[0]
            row_num +=1
        else: 
            sales[row[0]] = row[1]
     
# save the file    
wb.save('customer_sales_formatted_2020_2024.xlsx')


# Load the Excel file
file_path = 'customer_sales_formatted_2020_2024.xlsx'
df = pd.read_excel(file_path)

df = df.dropna() # remove empty rows
df.to_excel("output.xlsx")


wb = Workbook()
ws = wb.active
ws.title = "ANOVA Results"

# Write headers to the results sheet
headers = ['Customer', 'F-statistic', 'p-value', 'Significant']
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Step 2: Iterate through the customers
row_num = 2  # Start from the second row (after the headers)

# Predefined list of months for comparison (for January to December)
month_list = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 
              'August', 'September', 'October', 'November', 'December']

# Initialize a dictionary to store sales data grouped by month for each customer
customer_monthly_sales = defaultdict(lambda: {month: [] for month in month_list})
monthly_sales = []
count = 0
# Step 1: Iterate through the rows of the DataFrame (each customer)
for _, row in df.iterrows():
    all_sales = []
    customer_name = row['Customer']
    sales_data = row[1:]  # The sales data starts from the second column onward
    # Step 2: Loop through each month column to group sales data by month
    for month in month_list:
        for column, sales_value in sales_data.items():
              # Step 3: Extract the month from the column header
            month_name = pd.to_datetime(column, errors='raise').strftime('%B')  # Month name (e.g., 'January')
            if month_name == month:
                monthly_sales.append(sales_value)
        all_sales.append(monthly_sales)
        monthly_sales = []
    
    f_stat, p_value = stats.f_oneway(*all_sales)
    significant = 'Yes' if p_value < 0.05 else 'No'
        
    ws.cell(row=row_num, column=1, value=customer_name)
    ws.cell(row=row_num, column=2, value=f_stat)
    ws.cell(row=row_num, column=3, value=f_stat)
    ws.cell(row=row_num, column=4, value=significant)
    
    row_num += 1

wb.save('customer_sales_anova_results_2020_2024.xlsx')